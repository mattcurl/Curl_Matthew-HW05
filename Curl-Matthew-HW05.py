# Matthew Curl
# Cosc 1010
# Lab Section: 13
# Submission Date: 11/19/2024
# Sources/ Help: Refrence Image: https://www.pinterest.com/pin/455496949807251379/ 
# I only did sonic's head and accidentally flipped the rows and columns so it is facing down

# make a program that will draw a pixel drawing in an excel sheet
# import openpyxl library and color and pattern fill and get column letter from open pyxl

import openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter
# use load in openpyxl to load the workbook
book = openpyxl.Workbook()
# open the active sheet
sheet = book.active
#set the column widths in range 1 to 17 to 3
for column in range(1,17):
    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = 3
# make a dictionary to hold colors and coords
colors = {
    "000000":['A3','A4','A5','A6','A7','A8','A9','A10','A11','A12','A13','A14','A18','A19','B2','B15','B16','B17','B19','C3','C7','C19','D4','D7','D19','E5','E7','E20','F4','F20','G3','G20','H2','H16','H19','I1','I2','I3','I4','I5','I6','I16','I19','J5','J16','J18','J19','J20','K4','K19','L3','L19','M2','M3','M4','M5','M6','M7','M8','M9','M10','M18','N11','N12','N13','N14','N15','N16','N17','N18'],
    "0F0CE8":['B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B18','C4','C5','C6','C9','C10','C11','C12','C13','C14','C15','C16','C17','C18','D5','D6','D10','D11','D12','D13','D14','D15','D16','D17','D18','E6','E9','E10','E11','E12','E13','E14','E15','E16','E17','E18','E19','F5','F6','F7','F8','F9','F10','F11','F12','F16','F17','F18','F19','G4','G5','G6','G7','G8','G9','G10','G11','G17','G18','G19','H3','H4','H5','H6','H7','H8','H9','H10','H11','H17','I7','I8','I9','I10','I11','I17','J6','J7','J8','J9','J10','J11','J12','K5','K6','K7','K8','K9','L4','L5','L6','L7','L8','L9'],
    "F8CB78":['C8','D8','D9','E8','K10','K11','K12','K17','K18','L12','L13','L14','L15','L16','L17'],
    "F6BD55":['L10','L11','L18','M11','M12','M13','M14','M15','M16','M17'],
    "FFFFFF":['F13','F14','F15','G12','G13','G14','G15','G16','H12','H13','H14','H15','H18','I12','I13','I14','I15','I18','J13','J14','J15','J17','K13','K14','K15','K16']
}
# make a for loop referencing the dictonary
for color, locations in colors.items():
    # set the color to paint to our color from the dictionary
    color_n = Color(rgb=color)
    # use pattern fill to fill each cell
    fill = PatternFill(patternType='solid',fgColor=color_n)
    # fill each spot in the sheet with its corresponding color
    for location in locations:
        sheet[location].fill = fill
# save the file and name it
book.save("Curl_Sheet4COSC.xlsx")